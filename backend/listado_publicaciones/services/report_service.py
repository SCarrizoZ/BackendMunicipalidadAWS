from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from reportlab.lib.utils import ImageReader
import matplotlib.pyplot as plt
from io import BytesIO
from django.db.models import Count, Q, F
from django.db.models.functions import TruncMonth
from datetime import datetime
import os
from django.conf import settings
import textwrap
import matplotlib

# Configuración de Matplotlib para evitar problemas de threading
matplotlib.use("Agg")

class ReportService:
    # Colores para gráficos
    CATEGORY_COLORS = {
        "Seguridad": "#FF6B6B",  # Rojo suave
        "Infraestructura": "#4ECDC4",  # Turquesa
        "Limpieza": "#45B7D1",  # Azul claro
        "Alumbrado": "#FFA07A",  # Salmón
        "Parques y Jardines": "#98FB98",  # Verde pálido
        "Tránsito": "#FFD700",  # Dorado
        "Otros": "#D3D3D3",  # Gris claro
    }

    MESES_ESPANOL = {
        1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun",
        7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic",
    }

    @staticmethod
    def generate_excel_report(publicaciones):
        """Genera un archivo Excel con las publicaciones dadas."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Publicaciones"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Encabezados
        headers = [
            "ID", "Título", "Categoría", "Junta Vecinal", "Fecha",
            "Situación", "Prioridad", "Descripción",
        ]
        ws.append(headers)

        # Aplicar estilos a los encabezados
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        # Agregar datos
        for pub in publicaciones:
            row = [
                pub.id,
                pub.titulo,
                pub.categoria.nombre,
                pub.junta_vecinal.nombre_junta,
                pub.fecha_publicacion.strftime("%Y-%m-%d"),
                pub.situacion.nombre if pub.situacion else "N/A",
                pub.prioridad,
                pub.descripcion,
            ]
            ws.append(row)

            # Aplicar bordes a las celdas de datos
            for cell in ws[ws.max_row]:
                cell.border = border

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        return wb

    @staticmethod
    def generate_pdf_report(publicaciones_filtradas, comentarios, departamento):
        """Genera un reporte PDF con gráficos y tablas."""
        buffer = BytesIO()
        
        # Generar los gráficos
        bar_chart_buffer = ReportService._generate_bar_chart(publicaciones_filtradas)
        pie_chart_buffer = ReportService._generate_pie_chart(publicaciones_filtradas)
        line_chart_buffer = ReportService._generate_line_chart(publicaciones_filtradas)
        table_buffer = ReportService._generate_tasa_resolucion_table(publicaciones_filtradas)

        # Crear el canvas de ReportLab
        pdf = canvas.Canvas(buffer, pagesize=letter)
        pdf.setTitle("Reporte de Publicaciones " + datetime.now().strftime("%d-%m-%Y"))
        width, height = letter

        # Agregar página de portada
        vertical_center = height / 2

        pdf.setFont("Helvetica-Bold", 60)
        pdf.drawCentredString(width / 2, vertical_center + 70, "Reporte")
        pdf.drawCentredString(width / 2, vertical_center, "de")
        pdf.drawCentredString(width / 2, vertical_center - 70, "Publicaciones")

        # Agregar logo
        logo_path = os.path.join(settings.BASE_DIR, "static", "images", "logo.png")
        if os.path.exists(logo_path):
            pdf.drawImage(
                logo_path, 50, height - 120, width=100, height=100, preserveAspectRatio=True
            )

        # Fecha
        pdf.setFont("Helvetica", 24)
        current_date = datetime.now()
        date_text = f"{ReportService.MESES_ESPANOL[current_date.month]} {current_date.year}"
        pdf.drawCentredString(width / 2, vertical_center - 150, date_text)

        # Subtítulos
        pdf.setFont("Helvetica", 14)
        pdf.drawCentredString(width / 2, vertical_center - 200, "Municipalidad de Calama")
        pdf.drawCentredString(
            width / 2, vertical_center - 220,
            ("Departamento: " + departamento) if departamento else "General",
        )

        # Líneas decorativas
        pdf.setStrokeColor(colors.yellow)
        pdf.setLineWidth(3)
        pdf.line(100, 50, width / 2, 50)
        pdf.setStrokeColor(colors.orange)
        pdf.line(width / 2, 50, width - 100, 50)

        pdf.showPage()

        # Página de Resumen
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(50, height - 40, "Municipalidad de Calama")

        if os.path.exists(logo_path):
            pdf.drawImage(logo_path, width - 100, height - 60, width=50, height=50, preserveAspectRatio=True)

        pdf.setFont("Helvetica-Bold", 16)
        pdf.drawCentredString(width / 2, height - 80, "Resumen")

        # Comentarios
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(50, height - 110, "Comentarios:")
        pdf.setFont("Helvetica", 12)

        y_position = height - 130
        max_width = 80

        if not comentarios:
            comentarios = "No hay comentarios."

        for linea in comentarios.split("\n"):
            for wrapped_line in textwrap.wrap(linea, max_width):
                if y_position < 70:
                    pdf.showPage()
                    y_position = height - 90
                    pdf.setFont("Helvetica", 12)

                pdf.drawString(50, y_position, wrapped_line.strip())
                y_position -= 20

        # Footer
        pdf.setFont("Helvetica", 8)
        pdf.drawString(50, 30, f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

        pdf.setStrokeColor(colors.yellow)
        pdf.setLineWidth(10)
        pdf.line(50, 20, width / 2, 20)
        pdf.setStrokeColor(colors.orange)
        pdf.line(width / 2, 20, width - 50, 20)

        pdf.showPage()

        # Gráficos (páginas horizontales)
        pdf.setPageSize((height, width))

        if bar_chart_buffer:
            pdf.setFont("Helvetica-Bold", 16)
            pdf.drawString(50, width - 50, "Gráfico de Barras")
            bar_chart_image = ImageReader(bar_chart_buffer)
            pdf.drawImage(bar_chart_image, 0, 0, width=height, height=width - 100, preserveAspectRatio=True)
            pdf.showPage()

        if pie_chart_buffer:
            pdf.setFont("Helvetica-Bold", 16)
            pdf.drawString(50, width - 50, "Gráfico Circular")
            pie_chart_image = ImageReader(pie_chart_buffer)
            pdf.drawImage(pie_chart_image, 0, 0, width=height, height=width - 100, preserveAspectRatio=True)
            pdf.showPage()

        if line_chart_buffer:
            pdf.setFont("Helvetica-Bold", 16)
            pdf.drawString(50, width - 50, "Gráfico de Líneas")
            line_chart_image = ImageReader(line_chart_buffer)
            pdf.drawImage(line_chart_image, 0, 0, width=height, height=width - 100, preserveAspectRatio=True)
            pdf.showPage()

        # Tabla
        if table_buffer:
            pdf.setFont("Helvetica-Bold", 16)
            title = "Tasa de Resolución por Departamento y Mes"
            pdf.drawString(50, width - 50, title)

            table_width = height * 0.9
            table_height = width - 100
            w, h = table_buffer.wrapOn(pdf, table_width, table_height)
            x = (height - w) / 2
            y = (width - 100 - h) / 2
            table_buffer.drawOn(pdf, x, y)
            pdf.showPage()

        pdf.save()
        buffer.seek(0)
        return buffer

    @staticmethod
    def _generate_bar_chart(publicaciones_filtradas):
        try:
            datos = (
                publicaciones_filtradas.annotate(mes=TruncMonth("fecha_publicacion"))
                .values("mes", "categoria__nombre")
                .annotate(total=Count("id"))
                .order_by("mes")
            )

            meses_dict = {}
            for dato in datos:
                mes_nombre = ReportService.MESES_ESPANOL[dato["mes"].month]
                if mes_nombre not in meses_dict:
                    meses_dict[mes_nombre] = {"name": mes_nombre}
                meses_dict[mes_nombre][dato["categoria__nombre"]] = dato["total"]

            data = list(meses_dict.values())
            if not data:
                return None

            meses = [item["name"] for item in data]
            categorias = list(set(key for item in data for key in item.keys() if key != "name"))
            valores_por_categoria = {categoria: [] for categoria in categorias}

            for item in data:
                for categoria in categorias:
                    valores_por_categoria[categoria].append(item.get(categoria, 0))

            plt.figure(figsize=(10, 6))
            bottom_stack = [0] * len(meses)

            for categoria, valores in valores_por_categoria.items():
                plt.bar(
                    meses, valores, label=categoria, bottom=bottom_stack,
                    color=ReportService.CATEGORY_COLORS.get(categoria, "#778899"),
                )
                bottom_stack = [i + j for i, j in zip(bottom_stack, valores)]

            plt.xlabel("Meses")
            plt.ylabel("Cantidad")
            plt.title("Publicaciones por Mes y Categoría")
            plt.legend(loc="upper left")
            plt.tight_layout()

            buffer = BytesIO()
            plt.savefig(buffer, format="png")
            buffer.seek(0)
            plt.close()
            return buffer
        except Exception as e:
            print(e)
            return None

    @staticmethod
    def _generate_pie_chart(publicaciones_filtradas):
        try:
            datos = (
                publicaciones_filtradas.values("categoria__nombre")
                .annotate(total=Count("id"))
                .order_by("-total")
            )

            if not datos:
                return None

            respuesta = [{"name": dato["categoria__nombre"], "value": dato["total"]} for dato in datos]
            categorias = [item["name"] for item in respuesta]
            valores = [item["value"] for item in respuesta]
            colores = [ReportService.CATEGORY_COLORS.get(categoria, "#778899") for categoria in categorias]

            fig, ax = plt.subplots(figsize=(10, 10))
            wedges, texts, autotexts = ax.pie(
                valores, labels=None, colors=colores, autopct="%1.1f%%", startangle=140,
                textprops={"color": "white", "fontsize": 14, "weight": "bold"},
            )
            ax.set_title("Distribución de Publicaciones por Categoría", fontsize=16)
            ax.legend(
                loc="center left", bbox_to_anchor=(1, 0.5),
                labels=[f"{c} ({v})" for c, v in zip(categorias, valores)], fontsize=10,
            )

            buffer = BytesIO()
            plt.tight_layout()
            plt.savefig(buffer, format="png", bbox_inches="tight")
            buffer.seek(0)
            plt.close()
            return buffer
        except Exception as e:
            print(e)
            return None

    @staticmethod
    def _generate_line_chart(publicaciones_filtradas):
        try:
            publicaciones_por_mes = (
                publicaciones_filtradas.annotate(mes=TruncMonth("fecha_publicacion"))
                .values("mes")
                .annotate(
                    recibidos=Count("id", filter=Q(situacion__nombre="Recibido")),
                    resueltos=Count("id", filter=Q(situacion__nombre="Resuelto")),
                    en_curso=Count("id", filter=Q(situacion__nombre="En curso")),
                )
                .order_by("mes")
            )

            if not publicaciones_por_mes:
                return None

            respuesta = []
            for dato in publicaciones_por_mes:
                mes = dato["mes"]
                respuesta.append({
                    "name": ReportService.MESES_ESPANOL[mes.month],
                    "recibidos": dato["recibidos"],
                    "resueltos": dato["resueltos"],
                    "en_curso": dato["en_curso"],
                })

            meses = [item["name"] for item in respuesta]
            recibidos = [item["recibidos"] for item in respuesta]
            resueltos = [item["resueltos"] for item in respuesta]
            en_curso = [item["en_curso"] for item in respuesta]

            plt.figure(figsize=(10, 6))
            plt.plot(meses, recibidos, label="Recibidos", marker="o", color="#82ca9d", linewidth=3)
            plt.plot(meses, resueltos, label="Resueltos", marker="o", color="#8884d8", linewidth=3)
            plt.plot(meses, en_curso, label="En curso", marker="o", color="#ff8042", linewidth=3)

            plt.xlabel("Meses")
            plt.ylabel("Cantidad")
            plt.title("Resoluciones por Mes")
            plt.legend(loc="upper left")
            plt.tight_layout()

            buffer = BytesIO()
            plt.savefig(buffer, format="png")
            buffer.seek(0)
            plt.close()
            return buffer
        except Exception as e:
            print(e)
            return None

    @staticmethod
    def _generate_tasa_resolucion_table(publicaciones_filtradas):
        try:
            datos = (
                publicaciones_filtradas.annotate(
                    mes=TruncMonth("fecha_publicacion"),
                    departamento_nombre=F("departamento__nombre"),
                )
                .values("mes", "departamento_nombre")
                .annotate(
                    total=Count("id"),
                    resueltos=Count("id", filter=Q(situacion__nombre="Resuelto")),
                )
                .order_by("mes", "departamento_nombre")
            )

            if not datos:
                return None

            encabezados = ["Departamento", "Mes", "Total", "Resueltos", "Tasa de Resolución"]
            filas = []
            for dato in datos:
                mes_nombre = ReportService.MESES_ESPANOL[dato["mes"].month]
                depto = dato["departamento_nombre"]
                total = dato["total"]
                resueltos = dato["resueltos"]
                tasa_resolucion = round(resueltos / total, 2) if total > 0 else 0
                filas.append([depto, mes_nombre, total, resueltos, f"{tasa_resolucion*100:.2f}%"])

            width, height = letter
            total_width = width * 0.9
            col_widths = [
                total_width * 0.25, total_width * 0.15, total_width * 0.15,
                total_width * 0.15, total_width * 0.30,
            ]

            tabla = Table([encabezados] + filas, colWidths=col_widths)
            tabla.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.orange),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 12),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
            ]))
            return tabla
        except Exception as e:
            print(e)
            return None
