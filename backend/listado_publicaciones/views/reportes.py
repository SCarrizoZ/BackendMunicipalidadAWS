from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from reportlab.lib.utils import ImageReader
import matplotlib.pyplot as plt
from io import BytesIO
from django.db.models import Count, Q, F
from django.db.models.functions import TruncMonth
from rest_framework.decorators import api_view, permission_classes
from listado_publicaciones.permissions import IsAdmin
from ..models import Publicacion
from ..filters import PublicacionFilter
from .auditoria import crear_auditoria
from datetime import datetime
import os
from django.conf import settings
import textwrap

# Configuración de Matplotlib para evitar problemas de threading
import matplotlib
matplotlib.use("Agg")

# Colores para gráficos
category_colors = {
    "Seguridad": "#FF6B6B",  # Rojo suave
    "Infraestructura": "#4ECDC4",  # Turquesa
    "Limpieza": "#45B7D1",  # Azul claro
    "Alumbrado": "#FFA07A",  # Salmón
    "Parques y Jardines": "#98FB98",  # Verde pálido
    "Tránsito": "#FFD700",  # Dorado
    "Otros": "#D3D3D3",  # Gris claro
}

meses_espanol = {
    1: "Ene",
    2: "Feb",
    3: "Mar",
    4: "Abr",
    5: "May",
    6: "Jun",
    7: "Jul",
    8: "Ago",
    9: "Sep",
    10: "Oct",
    11: "Nov",
    12: "Dic",
}

@api_view(["GET"])
@permission_classes([IsAdmin])
def export_to_excel(request):
    try:
        # Aplicar filtros usando PublicacionFilter
        filterset = PublicacionFilter(request.GET, queryset=Publicacion.objects.all())
        if not filterset.is_valid():
            return HttpResponse("Errores en los filtros", status=400)

        publicaciones = filterset.qs

        # Crear el libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Publicaciones"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Encabezados
        headers = [
            "ID",
            "Título",
            "Categoría",
            "Junta Vecinal",
            "Fecha",
            "Situación",
            "Prioridad",
            "Descripción",
        ]
        ws.append(headers)

        # Aplicar estilos a los encabezados
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        # Agregar datos
        for pub in publicaciones:
            row = [
                pub.id,
                pub.titulo,
                pub.categoria.nombre,
                pub.junta_vecinal.nombre_junta,
                pub.fecha_publicacion.strftime("%Y-%m-%d"),
                pub.situacion.nombre if pub.situacion else "N/A",
                pub.prioridad,
                pub.descripcion,
            ]
            ws.append(row)

            # Aplicar bordes a las celdas de datos
            for cell in ws[ws.max_row]:
                cell.border = border

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="publicaciones.xlsx"'
        wb.save(response)

        # Crear descripción detallada de los filtros aplicados
        filtros_aplicados = []
        for key, value in request.GET.items():
            if value:
                filtros_aplicados.append(f"{key}: {value}")

        descripcion_detallada = f"Exportación de {publicaciones.count()} registros a Excel"
        if filtros_aplicados:
            descripcion_detallada += f" - Filtros: {', '.join(filtros_aplicados)}"

        # Registrar auditoría
        crear_auditoria(
            usuario=request.user,
            accion="READ",  # O una acción específica como EXPORT
            modulo="Reportes",
            descripcion=descripcion_detallada,
            es_exitoso=True,
        )

        return response

    except Exception as e:
        # Registrar error en auditoría
        crear_auditoria(
            usuario=request.user,
            accion="READ",
            modulo="Reportes",
            descripcion=f"Error al exportar a Excel: {str(e)}",
            es_exitoso=False,
        )
        return HttpResponse("Error al generar el archivo Excel", status=500)


def generate_bar_chart(publicaciones_filtradas):
    """
    Genera un gráfico de barras apiladas con los datos proporcionados.
    """
    try:
        datos = (
            publicaciones_filtradas.annotate(mes=TruncMonth("fecha_publicacion"))
            .values("mes", "categoria__nombre")
            .annotate(total=Count("id"))
            .order_by("mes")
        )

        # Formatear los datos
        meses_dict = {}
        for dato in datos:
            mes_nombre = meses_espanol[dato["mes"].month]  # Ene, Feb, Mar, etc.
            if mes_nombre not in meses_dict:
                meses_dict[mes_nombre] = {"name": mes_nombre}

            meses_dict[mes_nombre][dato["categoria__nombre"]] = dato["total"]

        data = list(meses_dict.values())

        if not data:
            return None

        meses = [item["name"] for item in data]
        categorias = list(
            set(key for item in data for key in item.keys() if key != "name")
        )
        valores_por_categoria = {categoria: [] for categoria in categorias}

        for item in data:
            for categoria in categorias:
                valores_por_categoria[categoria].append(item.get(categoria, 0))

        plt.figure(figsize=(10, 6))
        bottom_stack = [0] * len(meses)

        for categoria, valores in valores_por_categoria.items():
            plt.bar(
                meses,
                valores,
                label=categoria,
                bottom=bottom_stack,
                color=category_colors.get(categoria, "#778899"),
            )
            bottom_stack = [i + j for i, j in zip(bottom_stack, valores)]

        plt.xlabel("Meses")
        plt.ylabel("Cantidad")
        plt.title("Publicaciones por Mes y Categoría")
        plt.legend(loc="upper left")
        plt.tight_layout()

        buffer = BytesIO()
        plt.savefig(buffer, format="png")
        buffer.seek(0)
        plt.close()
        return buffer
    except Exception as e:
        print(e)
        return None


def generate_pie_chart(publicaciones_filtradas):
    """
    Genera un gráfico circular con colores personalizados.
    """
    try:
        # Agrupar por categoría y contar publicaciones
        datos = (
            publicaciones_filtradas.values("categoria__nombre")
            .annotate(total=Count("id"))
            .order_by("-total")
        )  # Ordenar por total en orden descendente

        if not datos:
            return None

        # Dar formato a los datos
        respuesta = [
            {"name": dato["categoria__nombre"], "value": dato["total"]}
            for dato in datos
        ]
        categorias = [item["name"] for item in respuesta]
        valores = [item["value"] for item in respuesta]
        colores = [
            category_colors.get(categoria, "#778899") for categoria in categorias
        ]

        # Ajustar el tamaño y diseño del gráfico
        fig, ax = plt.subplots(figsize=(10, 10))
        wedges, texts, autotexts = ax.pie(
            valores,
            labels=None,  # Ocultar las etiquetas en el gráfico principal
            colors=colores,
            autopct="%1.1f%%",
            startangle=140,
            textprops={
                "color": "white",
                "fontsize": 14,
                "weight": "bold",
            },  # Cambiar color y tamaño del texto
        )
        ax.set_title("Distribución de Publicaciones por Categoría", fontsize=16)

        # Agregar leyendas externas
        ax.legend(
            loc="center left",
            bbox_to_anchor=(1, 0.5),  # Posicionar leyendas fuera del gráfico
            labels=[f"{c} ({v})" for c, v in zip(categorias, valores)],
            fontsize=10,
        )

        buffer = BytesIO()
        plt.tight_layout()  # Asegurar que los elementos no se superpongan
        plt.savefig(buffer, format="png", bbox_inches="tight")  # Ajustar al contenido
        buffer.seek(0)
        plt.close()
        return buffer
    except Exception as e:
        print(e)
        return None


def generate_line_chart(publicaciones_filtradas):
    """
    Genera un gráfico de líneas con los datos proporcionados.
    Utilizar vista "ResolucionesPorMes" para obtener datos.
    """

    try:

        # Anotar publicaciones agrupadas por mes
        publicaciones_por_mes = (
            publicaciones_filtradas.annotate(mes=TruncMonth("fecha_publicacion"))
            .values("mes")
            .annotate(
                recibidos=Count("id", filter=Q(situacion__nombre="Recibido")),
                resueltos=Count("id", filter=Q(situacion__nombre="Resuelto")),
                en_curso=Count("id", filter=Q(situacion__nombre="En curso")),
            )
            .order_by("mes")
        )

        if not publicaciones_por_mes:
            return None

        # Convertir el formato para la respuesta
        respuesta = []
        for dato in publicaciones_por_mes:
            mes = dato["mes"]
            respuesta.append(
                {
                    "name": meses_espanol[mes.month],
                    "recibidos": dato["recibidos"],
                    "resueltos": dato["resueltos"],
                    "en_curso": dato["en_curso"],
                }
            )

        # Crear el gráfico de líneas
        meses = [item["name"] for item in respuesta]
        recibidos = [item["recibidos"] for item in respuesta]
        resueltos = [item["resueltos"] for item in respuesta]
        en_curso = [item["en_curso"] for item in respuesta]

        plt.figure(figsize=(10, 6))
        plt.plot(
            meses,
            recibidos,
            label="Recibidos",
            marker="o",
            color="#82ca9d",
            linewidth=3,
        )
        plt.plot(
            meses,
            resueltos,
            label="Resueltos",
            marker="o",
            color="#8884d8",
            linewidth=3,
        )
        plt.plot(
            meses, en_curso, label="En curso", marker="o", color="#ff8042", linewidth=3
        )

        plt.xlabel("Meses")
        plt.ylabel("Cantidad")
        plt.title("Resoluciones por Mes")
        plt.legend(loc="upper left")
        plt.tight_layout()

        buffer = BytesIO()
        plt.savefig(buffer, format="png")
        buffer.seek(0)
        plt.close()
        return buffer
    except Exception as e:
        print(e)
        return None


def generate_tasa_resolucion_table(publicaciones_filtradas):
    """
    Genera una tabla con la tasa de resolución por departamento y mes.
    Utilizar vista "TasaResolucionDepartamento" para obtener datos.
    """
    try:
        datos = (
            publicaciones_filtradas.annotate(
                mes=TruncMonth("fecha_publicacion"),
                departamento_nombre=F("departamento__nombre"),
            )
            .values("mes", "departamento_nombre")
            .annotate(
                total=Count("id"),
                resueltos=Count("id", filter=Q(situacion__nombre="Resuelto")),
            )
            .order_by("mes", "departamento_nombre")
        )

        if not datos:
            return None

        # Formatear datos para la tabla
        encabezados = [
            "Departamento",
            "Mes",
            "Total",
            "Resueltos",
            "Tasa de Resolución",
        ]
        filas = []
        # Calcular la tasa de resolución
        for dato in datos:
            mes_nombre = meses_espanol[dato["mes"].month]
            depto = dato["departamento_nombre"]
            total = dato["total"]
            resueltos = dato["resueltos"]
            tasa_resolucion = round(resueltos / total, 2) if total > 0 else 0
            filas.append(
                [depto, mes_nombre, total, resueltos, f"{tasa_resolucion*100:.2f}%"]
            )

        width, height = letter  # (612, 792) en puntos
        total_width = width * 0.9  # Usar el 90% del ancho de la página

        # Distribuir el ancho total entre las columnas (proporcionalmente)
        col_widths = [
            total_width * 0.25,  # Departamento (25%)
            total_width * 0.15,  # Mes (15%)
            total_width * 0.15,  # Total (15%)
            total_width * 0.15,  # Resueltos (15%)
            total_width * 0.30,  # Tasa de Resolución (30%)
        ]

        # Crear la tabla con los nuevos anchos
        tabla = Table([encabezados] + filas, colWidths=col_widths)
        tabla.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.orange),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, -1), 10),
                ]
            )
        )
        return tabla
    except Exception as e:
        print(e)
        return None


@api_view(["GET"])
@permission_classes([IsAdmin])
def generate_pdf_report(request):
    try:
        comentarios = request.GET.get(
            "comentarios", "No se proporcionaron comentarios."
        )
        departamento = request.GET.get("departamento_reporte", "")
        # Aplicar filtros usando PublicacionFilter
        filterset = PublicacionFilter(request.GET, queryset=Publicacion.objects.all())
        if not filterset.is_valid():
            return HttpResponse("Errores en los filtros", status=400)

        publicaciones_filtradas = filterset.qs

        if publicaciones_filtradas.count() == 0:
            return HttpResponse("No hay datos para generar el reporte", status=404)

        # Generar los gráficos
        bar_chart_buffer = generate_bar_chart(publicaciones_filtradas)
        pie_chart_buffer = generate_pie_chart(publicaciones_filtradas)
        line_chart_buffer = generate_line_chart(publicaciones_filtradas)
        table_buffer = generate_tasa_resolucion_table(publicaciones_filtradas)

        # Crear el PDF
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'inline; filename="reporte.pdf"'

        # Crear el canvas de ReportLab
        pdf = canvas.Canvas(response, pagesize=letter)

        pdf.setTitle("Reporte de Publicaciones " + datetime.now().strftime("%d-%m-%Y"))
        width, height = letter

        # Agregar página de portada
        # Calcular la posición vertical para centrar el texto
        vertical_center = height / 2

        pdf.setFont("Helvetica-Bold", 60)
        pdf.drawCentredString(width / 2, vertical_center + 70, "Reporte")
        pdf.drawCentredString(width / 2, vertical_center, "de")
        pdf.drawCentredString(width / 2, vertical_center - 70, "Publicaciones")

        # Agregar logo
        logo_path = os.path.join(
            settings.BASE_DIR,
            "static",
            "images",
            "logo.png",
        )  # Ajusta la ruta según tu estructura
        if os.path.exists(logo_path):
            pdf.drawImage(
                logo_path,
                50,  # Posición X (arriba a la izquierda)
                height - 120,  # Posición Y (arriba a la izquierda)
                width=100,  # Ancho del logo
                height=100,  # Alto del logo
                preserveAspectRatio=True,
            )

        # Fecha
        pdf.setFont("Helvetica", 24)
        current_date = datetime.now()
        date_text = f"{meses_espanol[current_date.month]} {current_date.year}"
        pdf.drawCentredString(width / 2, vertical_center - 150, date_text)

        # Subtítulos
        pdf.setFont("Helvetica", 14)
        pdf.drawCentredString(
            width / 2, vertical_center - 200, "Municipalidad de Calama"
        )
        pdf.drawCentredString(
            width / 2,
            vertical_center - 220,
            ("Departamento: " + departamento) if departamento else "General",
        )

        # Dibujar línea decorativa en la parte inferior
        pdf.setStrokeColor(colors.yellow)
        pdf.setLineWidth(3)
        pdf.line(100, 50, width / 2, 50)
        pdf.setStrokeColor(colors.orange)
        pdf.line(width / 2, 50, width - 100, 50)

        pdf.showPage()  # Finalizar la página de portada

        # Add header
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(50, height - 40, "Municipalidad de Calama")

        # Add logo (adjust path as needed)
        if os.path.exists(logo_path):
            pdf.drawImage(
                logo_path,
                width - 100,
                height - 60,
                width=50,
                height=50,
                preserveAspectRatio=True,
            )

        # Add title
        pdf.setFont("Helvetica-Bold", 16)
        pdf.drawCentredString(width / 2, height - 80, "Resumen")

        # Comentarios
        pdf.setFont("Helvetica-Bold", 14)  # Tamaño de letra menor
        pdf.drawString(50, height - 110, "Comentarios:")
        pdf.setFont("Helvetica", 12)

        pdf.setFont("Helvetica", 12)
        y_position = height - 130  # Posición inicial (debajo del encabezado)
        max_width = 80  # Máximo número de caracteres por línea

        if not comentarios:
            comentarios = "No hay comentarios."

        for linea in comentarios.split("\n"):
            # Dividir la línea en fragmentos que quepan en el ancho permitido
            for wrapped_line in textwrap.wrap(linea, max_width):
                if y_position < 70:  # Si se acerca al footer, crear nueva página
                    pdf.showPage()
                    y_position = height - 90  # Reiniciar posición en nueva página
                    pdf.setFont(
                        "Helvetica", 12
                    )  # Restablecer la fuente después del salto

                # Dibujar la línea en la posición actual
                pdf.drawString(50, y_position, wrapped_line.strip())
                y_position -= 20  # Espaciado entre líneas

        # Agregar footer en la última página
        pdf.setFont("Helvetica", 8)
        pdf.drawString(
            50, 30, f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        )

        # Draw colored lines at the bottom
        pdf.setStrokeColor(colors.yellow)
        pdf.setLineWidth(10)
        pdf.line(50, 20, width / 2, 20)
        pdf.setStrokeColor(colors.orange)
        pdf.line(width / 2, 20, width - 50, 20)

        pdf.showPage()

        # Rotar la página a horizontal
        pdf.setPageSize((height, width))

        # Agregar gráfico de barras al PDF en una página completa
        if bar_chart_buffer:
            pdf.setFont("Helvetica-Bold", 16)
            pdf.drawString(50, width - 50, "Gráfico de Barras")
            bar_chart_image = ImageReader(bar_chart_buffer)
            pdf.drawImage(
                bar_chart_image,
                0,
                0,
                width=height,
                height=width - 100,
                preserveAspectRatio=True,
            )
            pdf.showPage()

        # Rotar la página a horizontal
        pdf.setPageSize((height, width))

        # Agregar gráfico circular al PDF en una página completa
        if pie_chart_buffer:
            pdf.setFont("Helvetica-Bold", 16)
            pdf.drawString(50, width - 50, "Gráfico Circular")
            pie_chart_image = ImageReader(pie_chart_buffer)
            pdf.drawImage(
                pie_chart_image,
                0,
                0,
                width=height,
                height=width
                - 100,  # Mantener proporciones para el gráfico circular ajustado
                preserveAspectRatio=True,
            )
            pdf.showPage()

        # Rotar la página a horizontal
        pdf.setPageSize((height, width))

        if line_chart_buffer:
            pdf.setFont("Helvetica-Bold", 16)
            pdf.drawString(50, width - 50, "Gráfico de Líneas")

            line_chart_image = ImageReader(line_chart_buffer)
            pdf.drawImage(
                line_chart_image,
                0,
                0,
                width=height,
                height=width - 100,
                preserveAspectRatio=True,
            )
            pdf.showPage()

        # Rotar la página a horizontal
        pdf.setPageSize((height, width))

        # Tabla con la tasa de resolución por departamento y mes
        if table_buffer:
            pdf.setFont("Helvetica-Bold", 16)
            title = "Tasa de Resolución por Departamento y Mes"
            pdf.drawString(50, width - 50, title)

            # Calcular dimensiones de la tabla
            table_width = height * 0.9  # 90% del ancho de la página
            table_height = width - 100  # Alto total menos espacio para título

            # Obtener el alto real de la tabla
            w, h = table_buffer.wrapOn(pdf, table_width, table_height)

            # Calcular posición para centrar vertical y horizontalmente
            x = (height - w) / 2  # Centrar horizontalmente
            y = (width - 100 - h) / 2  # Centrar verticalmente en el espacio disponible

            # Dibujar la tabla en la posición calculada
            table_buffer.drawOn(pdf, x, y)
            pdf.showPage()

        # Finalizar el PDF
        pdf.save()

        # Crear descripción detallada de los filtros aplicados
        filtros_aplicados = []
        if departamento:
            filtros_aplicados.append(f"Departamento: {departamento}")

        # Obtener otros filtros de la query string
        filtros_query = []
        for key, value in request.GET.items():
            if key not in ["comentarios", "departamento_reporte"] and value:
                filtros_query.append(f"{key}: {value}")

        if filtros_query:
            filtros_aplicados.extend(filtros_query)

        descripcion_detallada = f"Generación de reporte PDF de publicaciones ({publicaciones_filtradas.count()} registros)"
        if filtros_aplicados:
            descripcion_detallada += f" - Filtros: {', '.join(filtros_aplicados)}"

        crear_auditoria(
            usuario=request.user,
            accion="GENERAR_REPORTE_PDF",
            modulo="Reportes",
            descripcion=descripcion_detallada,
            es_exitoso=True,
        )

        return response
    except Exception as e:
        # Registrar error en auditoría
        crear_auditoria(
            usuario=request.user,
            accion="GENERAR_REPORTE_PDF",
            modulo="Reportes",
            descripcion=f"Error al generar reporte PDF: {str(e)}",
            es_exitoso=False,
        )
        print(e)
        return HttpResponse("Error al generar el PDF", status=500)
